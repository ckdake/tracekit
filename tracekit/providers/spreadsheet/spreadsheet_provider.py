"""Spreadsheet provider for tracekit.

This module defines the SpreadsheetProvider class, which provides an interface
for interacting with activity data stored in local spreadsheet files.
"""

import decimal
from datetime import UTC, date, datetime, timedelta
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import openpyxl
from peewee import DoesNotExist

from tracekit.provider_sync import ProviderSync
from tracekit.providers.base_provider import FitnessProvider
from tracekit.providers.spreadsheet.spreadsheet_activity import SpreadsheetActivity
from tracekit.user_context import get_user_id


class SpreadsheetProvider(FitnessProvider):
    def __init__(self, path: str, config: dict[str, Any] | None = None):
        self.config = config or {}
        super().__init__(config)
        self.path = path

    @property
    def provider_name(self) -> str:
        """Return the name of this provider."""
        return "spreadsheet"

    @staticmethod
    def _seconds_to_hms(seconds: float | None) -> str:
        if seconds is None:
            return ""
        try:
            seconds = round(seconds)
            return str(timedelta(seconds=seconds))
        except Exception:
            return ""

    @staticmethod
    def _hms_to_seconds(hms: str | int | float | None) -> float | None:
        if not hms:
            return None
        # Accept numeric types directly
        if isinstance(hms, (int, float)):
            return float(hms)
        if isinstance(hms, decimal.Decimal):
            return float(hms)

        # Accept string types only
        if not isinstance(hms, str):
            return None
        # Ignore openpyxl types that are not string or numeric
        if hasattr(hms, "value") or hasattr(hms, "is_date"):
            return None
        try:
            t = datetime.strptime(hms, "%H:%M:%S")
            return t.hour * 3600 + t.minute * 60 + t.second
        except Exception:
            try:
                # Try MM:SS
                t = datetime.strptime(hms, "%M:%S")
                return t.minute * 60 + t.second
            except Exception:
                return None

    @staticmethod
    def _convert_to_gmt_timestamp(dt_val: str | datetime | date, source_tz: str) -> int:
        """Convert a date/datetime/str to a GMT Unix timestamp, assuming local time in source_tz."""
        tz = ZoneInfo(source_tz)

        if isinstance(dt_val, str):
            dt = datetime.fromisoformat(dt_val)
        elif isinstance(dt_val, datetime):
            dt = dt_val
        elif isinstance(dt_val, date):
            dt = datetime(dt_val.year, dt_val.month, dt_val.day)
        else:
            raise TypeError(f"Unsupported type for dt_val: {type(dt_val)}")

        dt = dt.replace(tzinfo=tz)
        return int(dt.astimezone(UTC).timestamp())

    def _pull_all_activities(self) -> list[SpreadsheetActivity]:
        xlsx_file = Path(self.path)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active

        if sheet is None:
            print(f"No active sheet found in {xlsx_file}")
            return []

        print(f"Processing spreadsheet file: {xlsx_file}")

        processed_count = 0

        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                continue  # Skip header row

            excel_row_number = i + 1  # Convert enumerate index to Excel row number (1-based)
            existing_activity = SpreadsheetActivity.get_or_none(
                (SpreadsheetActivity.spreadsheet_id == excel_row_number)
                & (SpreadsheetActivity.user_id == get_user_id())
            )
            if existing_activity is None:
                activity = self._process_parsed_data(
                    {
                        "file_name": str(xlsx_file),
                        "spreadsheet_id": excel_row_number,
                        "row": row,
                    }
                )
                if activity:
                    processed_count += 1

        print(f"Processed {processed_count} new spreadsheet activities")

        return self._get_activities()

    def _get_activities(self, date_filter: str | None = None) -> list["SpreadsheetActivity"]:
        """Get SpreadsheetActivity objects for a specific month."""
        file_activities = []

        if date_filter:
            year, month = map(int, date_filter.split("-"))
            for activity in SpreadsheetActivity.select().where(SpreadsheetActivity.user_id == get_user_id()):
                if hasattr(activity, "start_time") and activity.start_time:
                    try:
                        # Convert timestamp to datetime for comparison
                        dt = datetime.fromtimestamp(int(activity.start_time))
                        if dt.year == year and dt.month == month:
                            file_activities.append(activity)
                    except (ValueError, TypeError):
                        continue
        else:
            file_activities = list(SpreadsheetActivity.select().where(SpreadsheetActivity.user_id == get_user_id()))

        return file_activities

    def _process_parsed_data(self, parsed_data: dict) -> SpreadsheetActivity | None:
        try:
            existing_activity = SpreadsheetActivity.get(
                SpreadsheetActivity.spreadsheet_id == parsed_data.get("spreadsheet_id"),
                SpreadsheetActivity.user_id == get_user_id(),
            )
            return existing_activity
        except DoesNotExist:
            pass

        activity_kwargs: dict[str, Any] = {}

        start_time = self._convert_to_gmt_timestamp(parsed_data["row"][0], self.config.get("home_timezone", "UTC"))
        activity_kwargs["start_time"] = start_time

        if parsed_data["row"][1]:
            activity_kwargs["activity_type"] = parsed_data["row"][1]
        if parsed_data["row"][2]:
            activity_kwargs["location_name"] = parsed_data["row"][2]
        if parsed_data["row"][3]:
            activity_kwargs["city"] = parsed_data["row"][3]
        if parsed_data["row"][4]:
            activity_kwargs["state"] = parsed_data["row"][4]
        if parsed_data["row"][5]:
            activity_kwargs["temperature"] = parsed_data["row"][5]
        if parsed_data["row"][6]:
            activity_kwargs["equipment"] = parsed_data["row"][6]
        # duration_hms is stored as HH:MM:SS, but we want .duration in seconds
        if parsed_data["row"][7]:
            try:
                # Store the original HH:MM:SS format
                activity_kwargs["duration_hms"] = str(parsed_data["row"][7])
                # Also convert to seconds for the duration field
                activity_kwargs["duration"] = self._hms_to_seconds(str(parsed_data["row"][7]))
            except (ValueError, TypeError):
                pass
        if parsed_data["row"][8]:
            activity_kwargs["distance"] = parsed_data["row"][8]
        if parsed_data["row"][9]:
            activity_kwargs["max_speed"] = parsed_data["row"][9]
        if parsed_data["row"][10]:
            activity_kwargs["avg_heart_rate"] = parsed_data["row"][10]
        if parsed_data["row"][11]:
            activity_kwargs["max_heart_rate"] = parsed_data["row"][11]
        if parsed_data["row"][12]:
            activity_kwargs["calories"] = parsed_data["row"][12]
        if parsed_data["row"][13]:
            activity_kwargs["max_elevation"] = parsed_data["row"][13]
        if parsed_data["row"][14]:
            activity_kwargs["total_elevation_gain"] = parsed_data["row"][14]
        if parsed_data["row"][15]:
            activity_kwargs["with_names"] = parsed_data["row"][15]
        if parsed_data["row"][16]:
            activity_kwargs["avg_cadence"] = parsed_data["row"][16]
        if parsed_data["row"][17]:
            activity_kwargs["strava_id"] = parsed_data["row"][17]
        if parsed_data["row"][18]:
            activity_kwargs["garmin_id"] = parsed_data["row"][18]
        if parsed_data["row"][19]:
            activity_kwargs["ridewithgps_id"] = parsed_data["row"][19]
        if parsed_data["row"][20]:
            activity_kwargs["notes"] = parsed_data["row"][20]
            # Use notes field as name if not empty
            activity_kwargs["name"] = parsed_data["row"][20]
        activity_kwargs["source_file"] = parsed_data.get("file_name", "")
        activity_kwargs["source_file_type"] = "spreadsheet"
        activity_kwargs["spreadsheet_id"] = parsed_data["spreadsheet_id"]

        activity_kwargs["user_id"] = get_user_id()
        spreadsheet_activity = SpreadsheetActivity.create(**activity_kwargs)
        return spreadsheet_activity

    def _mark_all_months_as_synced(self) -> None:
        """Mark all months containing activities as synced for this provider."""
        # Get all unique months that have activities
        activities = SpreadsheetActivity.select().where(SpreadsheetActivity.user_id == get_user_id())
        unique_months = set()

        for activity in activities:
            if activity.start_time:
                try:
                    # Convert Unix timestamp to datetime and extract year-month
                    from datetime import datetime

                    dt = datetime.fromtimestamp(activity.start_time)
                    year_month = f"{dt.year:04d}-{dt.month:02d}"
                    unique_months.add(year_month)
                except (ValueError, TypeError):
                    continue

        # Create ProviderSync records for all months (if they don't already exist)
        for year_month in unique_months:
            existing_sync = ProviderSync.get_or_none(year_month, self.provider_name)
            if not existing_sync:
                ProviderSync.create(year_month=year_month, provider=self.provider_name, user_id=get_user_id())
                print(f"Marked {year_month} as synced for {self.provider_name}")

    def pull_activities(self, date_filter: str | None = None) -> list[SpreadsheetActivity]:
        """
        Process spreadsheet and return SpreadsheetActivity objects.
        If date_filter is None, returns all activities.
        If date_filter is provided (YYYY-MM), returns only activities from that month.
        """
        if date_filter is None:
            return self._pull_all_activities()

        existing_sync = ProviderSync.get_or_none(date_filter, self.provider_name)
        if not existing_sync:
            self._pull_all_activities()
            # For spreadsheet provider, mark ALL months containing activities as synced
            # since we process the entire spreadsheet regardless of date_filter
            self._mark_all_months_as_synced()
        else:
            print(f"Month {date_filter} already synced for {self.provider_name}")

        return self._get_activities(date_filter)

    def get_activity_by_id(self, activity_id: str) -> SpreadsheetActivity | None:
        """Get a specific activity by its spreadsheet_id."""
        try:
            return SpreadsheetActivity.get(
                SpreadsheetActivity.spreadsheet_id == int(activity_id),
                SpreadsheetActivity.user_id == get_user_id(),
            )
        except (ValueError, DoesNotExist):
            return None

    def update_activity(self, activity_data: dict[str, Any]) -> Any:
        """Update an existing SpreadsheetActivity with new data."""
        try:
            activity_id = activity_data.get("spreadsheet_id")
            if not activity_id:
                return None
            activity = SpreadsheetActivity.get(
                SpreadsheetActivity.spreadsheet_id == activity_id,
                SpreadsheetActivity.user_id == get_user_id(),
            )

            # Update the database record
            for key, value in activity_data.items():
                if key != "spreadsheet_id":  # Don't update the ID itself
                    setattr(activity, key, value)
            activity.save()

            # Also update the Excel file
            xlsx_file = Path(self.path)
            wb_obj = openpyxl.load_workbook(xlsx_file)
            sheet = wb_obj.active

            if sheet is None:
                return activity

            row_idx = int(activity_id)  # spreadsheet_id is now the Excel row number
            if row_idx <= 1 or row_idx > sheet.max_row:
                return activity

            # Update specific columns in the Excel file
            file_updated = False
            for key, value in activity_data.items():
                if key == "notes":
                    # Notes is in column 21 (1-based)
                    sheet.cell(row=row_idx, column=21, value=value)
                    file_updated = True
                elif key == "duration_hms":
                    # Duration is in column 8 (1-based)
                    sheet.cell(row=row_idx, column=8, value=value)
                    file_updated = True

            if file_updated:
                wb_obj.save(xlsx_file)

            return activity
        except Exception:
            return None

    def create_activity(self, activity_data: dict[str, Any]) -> str:
        """Create a new activity in the spreadsheet from activity data."""
        xlsx_file = Path(self.path)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active

        # Get the next row number
        next_row = sheet.max_row + 1

        # Prepare the row data for Excel
        row = [
            activity_data.get("start_time", ""),
            activity_data.get("activity_type", ""),
            activity_data.get("location_name", ""),
            activity_data.get("city", ""),
            activity_data.get("state", ""),
            activity_data.get("temperature", ""),
            activity_data.get("equipment", ""),
            activity_data.get("duration_hms", "") or self._seconds_to_hms(activity_data.get("duration")),
            activity_data.get("distance", ""),
            activity_data.get("max_speed", ""),
            activity_data.get("avg_heart_rate", ""),
            activity_data.get("max_heart_rate", ""),
            activity_data.get("calories", ""),
            activity_data.get("max_elevation", ""),
            activity_data.get("total_elevation_gain", ""),
            activity_data.get("with_names", ""),
            activity_data.get("avg_cadence", ""),
            activity_data.get("strava_id", ""),
            activity_data.get("garmin_id", ""),
            activity_data.get("ridewithgps_id", ""),
            activity_data.get("notes", ""),
        ]

        sheet.append(row)
        wb_obj.save(xlsx_file)

        # Also create a database record
        spreadsheet_id = str(next_row)

        # Prepare database record data
        db_data = activity_data.copy()
        db_data["spreadsheet_id"] = spreadsheet_id

        # Ensure duration_hms is set from either the provided value or converted from duration
        if not db_data.get("duration_hms") and db_data.get("duration"):
            db_data["duration_hms"] = self._seconds_to_hms(db_data["duration"])

        # Convert start_time if it's a string
        if isinstance(db_data.get("start_time"), str):
            db_data["start_time"] = self._convert_to_gmt_timestamp(
                db_data["start_time"], self.config.get("home_timezone", "UTC")
            )

        # Create the database record
        db_data["user_id"] = get_user_id()
        SpreadsheetActivity.create(**db_data)

        # Return the Excel row number as the spreadsheet_id
        return spreadsheet_id

    def get_all_gear(self) -> dict[str, str]:
        """Fetch gear/equipment from the spreadsheet."""
        xlsx_file = Path(self.path)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active

        if sheet is None:
            return {}

        gear_set = set()
        # The equipment column is index 6 (0-based)
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                continue  # Skip header row
            equipment = row[6] if len(row) > 6 else None
            if equipment:
                gear_set.add(str(equipment))
        # Use the equipment name as both key and value
        return {name: name for name in gear_set}

    def set_gear(self, gear_name: str, activity_id: str) -> bool:
        """Set the gear/equipment for a specific activity."""
        xlsx_file = Path(self.path)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active

        if sheet is None:
            return False

        row_idx = int(activity_id)
        if row_idx <= 1 or row_idx > sheet.max_row:
            return False

        # Equipment is in column 7 (1-based)
        sheet.cell(row=row_idx, column=7, value=gear_name)
        wb_obj.save(xlsx_file)
        return True

    def reset_activities(self, date_filter: str | None = None) -> int:
        """Reset (delete) Spreadsheet activities from local database."""
        from tracekit.providers.spreadsheet.spreadsheet_activity import (
            SpreadsheetActivity,
        )

        if date_filter:
            start_timestamp, end_timestamp = self._YYYY_MM_to_unixtime_range(
                date_filter, self.config.get("home_timezone", "US/Eastern")
            )

            return (
                SpreadsheetActivity.delete()
                .where(
                    (SpreadsheetActivity.start_time >= start_timestamp)
                    & (SpreadsheetActivity.start_time <= end_timestamp)
                    & (SpreadsheetActivity.user_id == get_user_id())
                )
                .execute()
            )
        else:
            return SpreadsheetActivity.delete().where(SpreadsheetActivity.user_id == get_user_id()).execute()
